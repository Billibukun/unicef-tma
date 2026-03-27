import csv
import json

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import models
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render

from common.models import Channel

from .forms import TrainingForm
from .models import Event, EventChannel, EventTeamMember, Training, TrainingCategory, TrainingCluster, TrainingLevel


@login_required
def event_list(request):
    qs = Event.objects.prefetch_related("states", "trainings", "event_channels__channel")
    user = request.user

    is_state_admin = user.groups.filter(name="State Admin").exists()
    if is_state_admin and user.state:
        qs = qs.filter(states=user.state)
    elif not user.is_superuser and not user.groups.filter(name__in=["UNICEF Admin", "National Admin"]).exists():
        # Training managers see only events with trainings they manage
        managed_training_ids = user.managed_trainings.values_list("pk", flat=True)
        qs = qs.filter(trainings__pk__in=managed_training_ids)

    return render(request, "trainings/event_list.html", {
        "events": qs.distinct(),
    })


@login_required
def event_detail(request, pk: int):
    event = get_object_or_404(
        Event.objects.prefetch_related("states", "event_channels__channel", "trainings__state"),
        pk=pk,
    )
    user = request.user

    trainings = event.trainings.select_related("state", "channel").order_by("state__name")
    if user.groups.filter(name="State Admin").exists() and user.state:
        trainings = trainings.filter(state=user.state)

    # Stats per training
    training_stats = []
    total_participants = 0
    total_cost = 0
    for t in trainings:
        p_count = t.participant_count
        cost = t.total_cost
        total_participants += p_count
        total_cost += cost
        training_stats.append({
            "training": t,
            "participants": p_count,
            "cost": cost,
        })

    return render(request, "trainings/event_detail.html", {
        "event": event,
        "training_stats": training_stats,
        "total_participants": total_participants,
        "total_cost": total_cost,
        "channels": event.event_channels.select_related("channel", "responsible_person"),
    })


@login_required
def event_create(request):
    from common.models import State

    if request.method == "POST":
        title = request.POST.get("title", "").strip()
        event_type = request.POST.get("event_type", "training")
        description = request.POST.get("description", "").strip()
        implementing_partner = request.POST.get("implementing_partner", "").strip()
        responsible_officer = request.POST.get("responsible_officer", "").strip()
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")
        state_ids = request.POST.getlist("states")
        channel_ids = request.POST.getlist("channels")

        if not title or not start_date or not end_date:
            messages.error(request, "Title and dates are required.")
            return redirect("event_create")

        event = Event.objects.create(
            title=title, event_type=event_type, description=description,
            implementing_partner=implementing_partner,
            responsible_officer=responsible_officer,
            start_date=start_date, end_date=end_date,
            created_by=request.user,
        )
        event.states.set(state_ids)

        # Create EventChannels
        for ch_id in channel_ids:
            payment = request.POST.get(f"payment_{ch_id}", "").strip()
            resp_name = request.POST.get(f"responsible_{ch_id}", "").strip()
            EventChannel.objects.create(
                event=event, channel_id=ch_id,
                payment_section=payment, responsible_name=resp_name,
            )

        # Auto-create a Training per state
        for state_id in state_ids:
            state = State.objects.get(pk=state_id)
            primary_channel_id = channel_ids[0] if channel_ids else None
            Training.objects.create(
                event=event,
                title=f"{title} — {state.name}",
                description=description,
                implementing_partner=implementing_partner,
                responsible_officer=responsible_officer,
                channel_id=primary_channel_id,
                state=state,
                destination=state.name,
                start_date=start_date,
                end_date=end_date,
                status="planned",
                created_by=request.user,
            )

        messages.success(request, f"Event '{title}' created with {len(state_ids)} state(s).")
        return redirect("event_detail", pk=event.pk)

    return render(request, "trainings/event_form.html", {
        "states": State.objects.all(),
        "channels": Channel.objects.filter(is_active=True),
    })


@login_required
def event_edit(request, pk: int):
    event = get_object_or_404(Event, pk=pk)
    from common.models import State

    if request.method == "POST":
        event.title = request.POST.get("title", "").strip()
        event.event_type = request.POST.get("event_type", "training")
        event.description = request.POST.get("description", "").strip()
        event.implementing_partner = request.POST.get("implementing_partner", "").strip()
        event.responsible_officer = request.POST.get("responsible_officer", "").strip()
        event.start_date = request.POST.get("start_date")
        event.end_date = request.POST.get("end_date")
        event.status = request.POST.get("status", event.status)
        event.save()
        messages.success(request, f"Event '{event.title}' updated.")
        return redirect("event_manage", pk=event.pk)

    return render(request, "trainings/event_form.html", {
        "event": event,
        "states": State.objects.all(),
        "channels": Channel.objects.filter(is_active=True),
        "selected_states": list(event.states.values_list("pk", flat=True)),
        "event_channels": {ec.channel_id: ec for ec in event.event_channels.all()},
    })


@login_required
def event_manage(request, pk: int):
    """Full-page Event Management Hub — configure all settings."""
    event = get_object_or_404(
        Event.objects.prefetch_related("states", "event_channels__channel", "trainings__state"),
        pk=pk,
    )
    from common.models import LGA, State, TrainingRole

    # Select which state's training to configure
    state_id = request.GET.get("state")
    trainings = event.trainings.select_related("state", "channel").order_by("state__name")
    training = None
    if state_id:
        training = trainings.filter(state_id=state_id).first()
    if not training:
        training = trainings.first()

    categories = training.categories.prefetch_related("levels") if training else []
    levels = training.levels.all() if training else []
    clusters = training.clusters.prefetch_related("lgas") if training else []

    # Available LGAs (not yet in any cluster)
    state_lgas = training.state.lgas.all() if training and training.state else LGA.objects.none()
    assigned_lga_ids = TrainingCluster.objects.filter(training=training).values_list("lgas", flat=True) if training else []
    available_lgas = state_lgas.exclude(pk__in=assigned_lga_ids)

    # Channels
    event_channels = event.event_channels.select_related("channel", "responsible_person")
    all_channels = Channel.objects.filter(is_active=True)

    # Managers for this training
    managers = training.managers.all() if training else []

    # Available roles
    available_roles = TrainingRole.objects.filter(
        models.Q(channel=training.channel) | models.Q(channel__isnull=True)
    ) if training else []

    manage_url = f"/trainings/events/{event.pk}/manage/"
    if training:
        manage_url += f"?state={training.state_id}"

    return render(request, "trainings/event_manage.html", {
        "event": event,
        "trainings": trainings,
        "training": training,
        "categories": categories,
        "levels": levels,
        "clusters": clusters,
        "state_lgas": state_lgas,
        "available_lgas": available_lgas,
        "event_channels": event_channels,
        "all_channels": all_channels,
        "managers": managers,
        "available_roles": available_roles,
        "manage_url": manage_url,
    })


@login_required
def event_search_users(request, pk: int):
    """HTMX: search users for team member / responsible person assignment."""
    from accounts.models import CustomUser
    q = request.GET.get("q", "").strip()
    target = request.GET.get("target", "team")  # team, responsible, state_admin
    results = []
    if len(q) >= 2:
        from django.db.models import Q
        results = CustomUser.objects.filter(
            Q(first_name__icontains=q) | Q(last_name__icontains=q) |
            Q(username__icontains=q) | Q(email__icontains=q),
            is_active=True,
        ).exclude(
            participant_profile__isnull=False  # exclude DMs
        )[:10]
    return render(request, "trainings/partials/user_search_results.html", {
        "results": results, "event_pk": pk, "target": target,
    })


@login_required
def event_add_team_member(request, pk: int):
    """Add a team member (lead or support) to an event."""
    event = get_object_or_404(Event, pk=pk)
    if request.method == "POST":
        user_id = request.POST.get("user_id")
        role = request.POST.get("role", "support")
        if user_id:
            from accounts.models import CustomUser
            user = get_object_or_404(CustomUser, pk=user_id)
            EventTeamMember.objects.get_or_create(
                event=event, user=user, defaults={"role": role}
            )
            messages.success(request, f"{user.get_full_name()} added as {role}.")
    return redirect("event_manage", pk=pk)


@login_required
def event_remove_team_member(request, pk: int, member_id: int):
    event = get_object_or_404(Event, pk=pk)
    if request.method == "POST":
        EventTeamMember.objects.filter(pk=member_id, event=event).delete()
    return redirect("event_manage", pk=pk)


@login_required
def event_update_channel(request, pk: int, channel_id: int):
    """Update implementing partner and responsible person for an EventChannel."""
    ec = get_object_or_404(EventChannel, pk=channel_id, event_id=pk)
    if request.method == "POST":
        ec.implementing_partner = request.POST.get("implementing_partner", "").strip()
        ec.payment_section = request.POST.get("payment_section", "").strip()
        ec.responsible_name = request.POST.get("responsible_name", "").strip()
        resp_id = request.POST.get("responsible_person") or None
        ec.responsible_person_id = resp_id
        ec.is_primary = "is_primary" in request.POST
        ec.save()
        messages.success(request, f"{ec.channel.name} channel updated.")
    return redirect("event_manage", pk=pk)


@login_required
def event_add_channel(request, pk: int):
    """Add a channel to an event."""
    event = get_object_or_404(Event, pk=pk)
    if request.method == "POST":
        channel_id = request.POST.get("channel")
        if channel_id:
            EventChannel.objects.get_or_create(
                event=event, channel_id=channel_id,
                defaults={
                    "implementing_partner": request.POST.get("implementing_partner", ""),
                    "payment_section": request.POST.get("payment_section", ""),
                }
            )
    return redirect("event_manage", pk=pk)


@login_required
def event_assign_state_admin(request, pk: int):
    """Assign a state admin to a training under this event + send email."""
    from django.conf import settings as s
    from django.core.mail import send_mail
    from accounts.models import CustomUser

    event = get_object_or_404(Event, pk=pk)
    if request.method == "POST":
        user_id = request.POST.get("user_id")
        training_id = request.POST.get("training_id")
        if user_id and training_id:
            user = get_object_or_404(CustomUser, pk=user_id)
            training = get_object_or_404(Training, pk=training_id, event=event)
            training.managers.add(user)

            # Send notification email
            if user.email:
                base = getattr(s, "BASE_URL", "https://tma.worksiapps.com")
                try:
                    send_mail(
                        f"UNICEF TMA — You've been assigned to {event.title}",
                        f"Dear {user.get_full_name()},\n\n"
                        f"You have been assigned as a State Admin for:\n\n"
                        f"Event: {event.title}\n"
                        f"State: {training.state}\n"
                        f"Type: {event.get_event_type_display()}\n"
                        f"Dates: {event.start_date} — {event.end_date}\n\n"
                        f"Please log in to set up your state's operational areas:\n"
                        f"{base}\n\n"
                        f"---\nUNICEF Training Management Application",
                        None, [user.email], fail_silently=True,
                    )
                except Exception:
                    pass

            messages.success(request, f"{user.get_full_name()} assigned to {training.state}.")
    return redirect("event_manage", pk=pk)


def _settings_redirect(training, request=None):
    """Redirect back to training detail or custom next URL."""
    from django.http import HttpResponseRedirect
    from django.urls import reverse
    if request:
        next_url = request.POST.get("next", "")
        if next_url:
            return HttpResponseRedirect(next_url)
    return HttpResponseRedirect(reverse("training_detail", args=[training.pk]) + "?settings=1")


@login_required
def training_list(request):
    qs = Training.objects.select_related("channel", "created_by", "state")

    # Role-based filtering
    user = request.user
    is_state_admin = user.groups.filter(name="State Admin").exists()
    is_training_manager = user.managed_trainings.exists()
    if is_state_admin and user.state:
        qs = qs.filter(state=user.state)
    elif is_training_manager and not user.is_superuser:
        if not user.groups.filter(name__in=["UNICEF Admin", "National Admin", "UNICEF HQ"]).exists():
            qs = user.managed_trainings.all().select_related("channel", "created_by", "state")

    # Filters from GET params
    channel_id = request.GET.get("channel")
    status = request.GET.get("status")
    state = request.GET.get("state")

    if channel_id:
        qs = qs.filter(channel_id=channel_id)
    if status:
        qs = qs.filter(status=status)
    if state:
        qs = qs.filter(state_id=state)

    from django.core.paginator import Paginator
    from common.models import State
    per_page = request.GET.get("per_page", 15)
    paginator = Paginator(qs, per_page)
    page = request.GET.get("page")
    trainings_page = paginator.get_page(page)

    context = {
        "trainings": trainings_page,
        "states": State.objects.all(),
        "channels": Channel.objects.filter(is_active=True),
        "status_choices": Training.STATUS_CHOICES,
        "current_channel": channel_id or "",
        "current_status": status or "",
        "current_state": state or "",
    }
    return render(request, "trainings/training_list.html", context)


@login_required
def training_create(request):
    if request.method == "POST":
        form = TrainingForm(request.POST)
        if form.is_valid():
            training = form.save(commit=False)
            training.created_by = request.user
            training.save()
            messages.success(request, f'Training "{training.title}" created.')
            return redirect("training_detail", pk=training.pk)
    else:
        form = TrainingForm()

    return render(request, "trainings/training_form.html", {
        "form": form,
        "page_title": "Create Training",
    })


@login_required
def training_detail(request, pk: int):
    training = get_object_or_404(
        Training.objects.select_related("channel", "state", "created_by"),
        pk=pk,
    )
    levels = training.levels.all()
    categories = training.categories.prefetch_related("levels")
    clusters = training.clusters.prefetch_related("lgas")
    assignments = (
        training.assignments
        .select_related(
            "participant", "participant__channel", "participant__state",
            "role", "cluster", "training_category",
        )
        .order_by("training_category__name", "cluster__name", "participant__last_name")
    )

    # Device manager scoping — only see their cluster or LGA participants
    user_cluster = None
    user_lga = None
    is_device_mgr = hasattr(request.user, "participant_profile") and request.user.participant_profile is not None
    if is_device_mgr and not request.user.is_superuser and not request.user.groups.filter(name__in=["UNICEF Admin", "National Admin", "State Admin"]).exists():
        try:
            dm_participant = request.user.participant_profile
            dm_assignment = training.assignments.get(participant=dm_participant)
            user_cluster = dm_assignment.cluster
            user_lga = dm_participant.lga

            if user_cluster:
                # Filter by cluster
                assignments = assignments.filter(cluster=user_cluster)
            elif user_lga:
                # No cluster — filter by participant's LGA
                assignments = assignments.filter(participant__lga=user_lga)
        except Exception:
            pass

    # Filters
    cat_filter = request.GET.get("cat")
    cluster_filter = request.GET.get("cluster")
    role_filter = request.GET.get("role")
    lga_filter = request.GET.get("lga")
    channel_filter = request.GET.get("channel")
    if cat_filter:
        assignments = assignments.filter(training_category_id=cat_filter)
    if cluster_filter:
        assignments = assignments.filter(cluster_id=cluster_filter)
    if role_filter:
        assignments = assignments.filter(role_id=role_filter)
    if lga_filter:
        assignments = assignments.filter(participant__lga_id=lga_filter)
    if channel_filter:
        assignments = assignments.filter(participant__channel_id=channel_filter)

    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(assignments, 25)
    page = request.GET.get("page")
    assignments_page = paginator.get_page(page)

    # Financials summary
    financials = []
    for cat in categories:
        cat_assignments = cat.assignments.all()
        count = cat_assignments.count()
        cat_total = sum(a.total_payment for a in cat_assignments)
        financials.append({
            "category": cat,
            "count": count,
            "total": cat_total,
        })
    grand_total = sum(f["total"] for f in financials)

    from common.models import TrainingRole
    available_roles = TrainingRole.objects.filter(
        models.Q(channel=training.channel) | models.Q(channel__isnull=True)
    )

    # Scope devices — admins see all devices in the state, device managers see their LGA
    from devices.models import Device as DeviceModel
    if is_device_mgr:
        if user_cluster:
            cluster_lga_ids = user_cluster.lgas.values_list("pk", flat=True)
            training_devices = DeviceModel.objects.filter(state=training.state, lga_id__in=cluster_lga_ids)
        elif user_lga:
            training_devices = DeviceModel.objects.filter(state=training.state, lga=user_lga)
        else:
            training_devices = DeviceModel.objects.filter(training=training)
    else:
        # Admins see all devices in the training's state
        training_devices = DeviceModel.objects.filter(state=training.state).select_related("assigned_to", "lga")

    # Paginate banks tab
    bank_assignments = training.assignments.select_related(
        "participant", "participant__channel",
    ).order_by("participant__last_name")
    bank_paginator = Paginator(bank_assignments, 15)
    bank_page = bank_paginator.get_page(request.GET.get("bp"))

    # Paginate devices tab
    dev_paginator = Paginator(training_devices, 15)
    dev_page = dev_paginator.get_page(request.GET.get("dp"))

    # Cluster report — participant counts per cluster
    cluster_stats = []
    for cluster in clusters:
        cluster_lga_ids = cluster.lgas.values_list("pk", flat=True)
        cluster_assignments = training.assignments.filter(cluster=cluster)
        cluster_count = cluster_assignments.count()
        # Also count by category within cluster
        cat_counts = []
        for cat in categories:
            cat_count = cluster_assignments.filter(training_category=cat).count()
            cat_counts.append({"category": cat, "count": cat_count})
        lga_names = list(cluster.lgas.values_list("name", flat=True))
        cluster_stats.append({
            "cluster": cluster,
            "count": cluster_count,
            "lga_count": len(lga_names),
            "lgas": lga_names,
            "cat_counts": cat_counts,
        })
    # Unassigned — participants not in any cluster
    unassigned_count = training.assignments.filter(cluster__isnull=True).count()
    # Missing LGA
    no_lga_count = training.assignments.filter(participant__lga__isnull=True).count()

    # Duplicate detection — phone, email, bank account
    from collections import Counter
    from banks.models import BankAccount
    all_training_assignments = list(training.assignments.select_related(
        "participant", "participant__lga",
    ))
    phone_map: dict[str, list] = {}
    email_map: dict[str, list] = {}
    bank_map: dict[str, list] = {}
    for a in all_training_assignments:
        p = a.participant
        if p.phone:
            key = p.phone.strip().replace(" ", "").replace("+234", "0").lstrip("+")
            phone_map.setdefault(key, []).append(a)
        if p.email:
            key = p.email.strip().lower()
            email_map.setdefault(key, []).append(a)
        try:
            ba = p.bank_account
            if ba and ba.account_number:
                bank_map.setdefault(ba.account_number, []).append(a)
        except BankAccount.DoesNotExist:
            pass

    duplicates = []
    seen_pairs = set()
    for label, mapping, dtype in [
        ("Phone", phone_map, "phone"),
        ("Email", email_map, "email"),
        ("Bank Account", bank_map, "bank"),
    ]:
        for value, assigns in mapping.items():
            if len(assigns) > 1:
                ids = tuple(sorted(a.pk for a in assigns))
                if (dtype, ids) not in seen_pairs:
                    seen_pairs.add((dtype, ids))
                    duplicates.append({
                        "type": dtype,
                        "label": label,
                        "value": value,
                        "assignments": assigns,
                    })

    return render(request, "trainings/training_detail.html", {
        "training": training,
        "levels": levels,
        "categories": categories,
        "clusters": clusters,
        "cluster_stats": cluster_stats,
        "unassigned_cluster_count": unassigned_count,
        "no_lga_count": no_lga_count,
        "duplicates": duplicates,
        "device_managers": training.assignments.filter(
            training_category__is_device_manager=True,
        ).select_related(
            "participant", "participant__user", "participant__lga",
            "cluster", "training_category",
        ).order_by("cluster__name", "participant__last_name"),
        "assignments": assignments_page,
        "all_assignments": bank_page,
        "bank_page": bank_page,
        "training_devices": dev_page,
        "dev_page": dev_page,
        "total_assignments": paginator.count,
        "training_devices": training_devices,
        "user_cluster": user_cluster,
        "user_lga": user_lga,
        "is_device_mgr": is_device_mgr,
        "financials": financials,
        "grand_total": grand_total,
        "available_roles": available_roles,
        "state_lgas": training.state.lgas.all() if training.state else [],
        "available_lgas": training.state.lgas.exclude(
            pk__in=TrainingCluster.objects.filter(training=training).values_list("lgas", flat=True)
        ) if training.state else [],
        "cat_filter": cat_filter,
        "cluster_filter": cluster_filter,
        "role_filter": role_filter,
        "lga_filter": lga_filter,
        "channel_filter": channel_filter,
        "channels": Channel.objects.filter(is_active=True),
    })


@login_required
def training_edit(request, pk: int):
    training = get_object_or_404(Training, pk=pk)

    if request.method == "POST":
        form = TrainingForm(request.POST, instance=training)
        if form.is_valid():
            form.save()
            messages.success(request, f'Training "{training.title}" updated.')
            return redirect("training_detail", pk=training.pk)
    else:
        form = TrainingForm(instance=training)

    return render(request, "trainings/training_form.html", {
        "form": form,
        "training": training,
        "page_title": "Edit Training",
    })


@login_required
def training_delete(request, pk: int):
    training = get_object_or_404(Training, pk=pk)

    if request.method == "POST":
        title = training.title
        training.delete()
        messages.success(request, f'Training "{title}" deleted.')
        return redirect("training_list")

    return render(request, "trainings/training_confirm_delete.html", {
        "training": training,
    })


@login_required
def add_category(request, pk: int):
    """HTMX: add a category to a training."""
    training = get_object_or_404(Training, pk=pk)
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        dsa_rate = request.POST.get("dsa_rate", 0)
        local_transport = request.POST.get("local_transport", 0)
        travel_mode = "enabled" if "travel_enabled" in request.POST else "none"
        expires = request.POST.get("registration_expires", "")

        level_ids = request.POST.getlist("levels")

        arrival_day = "arrival_day" in request.POST
        collect_nin = "collect_nin" in request.POST
        is_device_manager = "is_device_manager" in request.POST

        if name:
            from django.utils.dateparse import parse_datetime
            cat = TrainingCategory.objects.create(
                training=training,
                name=name,
                arrival_day=arrival_day,
                collect_nin=collect_nin,
                is_device_manager=is_device_manager,
                dsa_rate=dsa_rate or 0,
                local_transport=local_transport or 0,
                travel_mode=travel_mode,
                registration_expires=parse_datetime(expires) if expires else None,
            )
            if level_ids:
                cat.levels.set(level_ids)

    return _settings_redirect(training, request)


@login_required
def add_cluster(request, pk: int):
    training = get_object_or_404(Training, pk=pk)
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        training_centre = request.POST.get("training_centre", "").strip()
        all_lgas = "all_lgas" in request.POST
        lga_ids = request.POST.getlist("lgas")
        if name:
            cluster, _ = TrainingCluster.objects.get_or_create(
                training=training, name=name,
                defaults={"training_centre": training_centre},
            )
            if all_lgas and training.state:
                assigned = TrainingCluster.objects.filter(training=training).exclude(pk=cluster.pk).values_list("lgas", flat=True)
                remaining = training.state.lgas.exclude(pk__in=assigned)
                cluster.lgas.set(remaining)
            elif lga_ids:
                cluster.lgas.set(lga_ids)
    return _settings_redirect(training, request)


@login_required
def toggle_category_registration(request, cat_id: int):
    cat = get_object_or_404(TrainingCategory.objects.select_related("training"), pk=cat_id)
    if request.method == "POST":
        cat.registration_open = not cat.registration_open
        if cat.registration_open:
            # Clear expiry when reopening so it doesn't block
            cat.registration_expires = None
        cat.save(update_fields=["registration_open", "registration_expires"])
    return _settings_redirect(cat.training, request)


@login_required
def edit_category(request, cat_id: int):
    cat = get_object_or_404(TrainingCategory.objects.select_related("training"), pk=cat_id)
    if request.method == "POST":
        cat.name = request.POST.get("name", cat.name).strip()
        cat.dsa_rate = request.POST.get("dsa_rate", cat.dsa_rate) or 0
        cat.local_transport = request.POST.get("local_transport", cat.local_transport) or 0
        cat.travel_mode = "enabled" if "travel_enabled" in request.POST else "none"
        cat.arrival_day = "arrival_day" in request.POST
        cat.collect_nin = "collect_nin" in request.POST
        cat.is_device_manager = "is_device_manager" in request.POST
        cat.device_target_category_id = request.POST.get("device_target_category") or None
        expires = request.POST.get("registration_expires", "")
        if expires:
            from django.utils.dateparse import parse_datetime
            cat.registration_expires = parse_datetime(expires)
        cat.save()
        level_ids = request.POST.getlist("levels")
        cat.levels.set(level_ids)
    # Regular redirect — modal will reopen via ?settings=1
    from django.http import HttpResponseRedirect
    from django.urls import reverse
    return HttpResponseRedirect(reverse("training_detail", args=[cat.training_id]) + "?settings=1")


@login_required
def delete_category(request, cat_id: int):
    cat = get_object_or_404(TrainingCategory.objects.select_related("training"), pk=cat_id)
    training = cat.training
    if request.method == "POST":
        cat.delete()
    return _settings_redirect(training, request)


@login_required
def edit_cluster(request, cluster_id: int):
    cluster = get_object_or_404(TrainingCluster.objects.select_related("training"), pk=cluster_id)
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        training_centre = request.POST.get("training_centre", "").strip()
        if name:
            cluster.name = name
        cluster.training_centre = training_centre
        cluster.save()
        lga_ids = request.POST.getlist("lgas")
        cluster.lgas.set(lga_ids)
    return _settings_redirect(cluster.training, request)


@login_required
def assignment_detail(request, assignment_id: int):
    """Training-specific participant detail — shows only data for this training."""
    from trainings.models import TrainingAssignment
    assignment = get_object_or_404(
        TrainingAssignment.objects.select_related(
            "training", "training__state", "participant", "participant__channel",
            "participant__state", "participant__lga", "training_category", "cluster", "role",
        ),
        pk=assignment_id,
    )
    participant = assignment.participant
    training = assignment.training

    # Bank account
    bank_account = None
    try:
        bank_account = participant.bank_account
    except Exception:
        pass

    # Devices for this training only
    from devices.models import Device
    devices = Device.objects.filter(
        assigned_to=participant, training=training,
    ).select_related("state")

    return render(request, "trainings/assignment_detail.html", {
        "assignment": assignment,
        "participant": participant,
        "training": training,
        "bank_account": bank_account,
        "devices": devices,
    })


@login_required
def edit_assignment(request, assignment_id: int):
    """Edit training-specific data for a participant (role, category, cluster)."""
    from trainings.models import TrainingAssignment
    from common.models import TrainingRole
    assignment = get_object_or_404(
        TrainingAssignment.objects.select_related("training", "participant", "training_category", "cluster", "role"),
        pk=assignment_id,
    )
    training = assignment.training

    if request.method == "POST":
        assignment.training_category_id = request.POST.get("category") or None

        # Auto-assign cluster from participant LGA
        if assignment.participant.lga:
            cluster = TrainingCluster.objects.filter(
                training=training, lgas=assignment.participant.lga
            ).first()
            if cluster:
                assignment.cluster = cluster

        # Attendance
        att = request.POST.get("attended")
        if att == "1":
            assignment.attended = True
        elif att == "0":
            assignment.attended = False
        else:
            assignment.attended = None
        assignment.attended_marked_by = request.user

        # Travel
        assignment.outbound_mode = request.POST.get("outbound_mode", "none")
        assignment.outbound_from_id = request.POST.get("outbound_from") or None
        assignment.outbound_to_id = training.state_id
        assignment.return_mode = request.POST.get("return_mode", "none")
        assignment.return_from_id = training.state_id
        assignment.return_to_id = request.POST.get("return_to") or None
        assignment.outbound_air_claim = request.POST.get("outbound_air_claim") or 0
        assignment.return_air_claim = request.POST.get("return_air_claim") or 0
        assignment.outbound_airline = request.POST.get("outbound_airline", "").strip()
        assignment.return_airline = request.POST.get("return_airline", "").strip()

        # Handle ticket uploads
        if request.FILES.get("outbound_ticket"):
            assignment.outbound_ticket = request.FILES["outbound_ticket"]
        if request.FILES.get("return_ticket"):
            assignment.return_ticket = request.FILES["return_ticket"]

        # Auto-calculate road mileage
        assignment.calculate_mileage()

        assignment.save()
        messages.success(request, f"Updated {assignment.participant.full_name}'s assignment.")
        return redirect("training_detail", pk=training.pk)

    from common.models import State
    categories = training.categories.all()
    clusters = training.clusters.all()
    states = State.objects.all()

    # Build category travel mode map for Alpine.js
    cat_travel = {str(c.pk): c.travel_mode for c in categories}

    return render(request, "trainings/edit_assignment.html", {
        "assignment": assignment,
        "categories": categories,
        "clusters": clusters,
        "states": states,
        "cat_travel_json": json.dumps(cat_travel),
    })


@login_required
def delete_cluster(request, cluster_id: int):
    cluster = get_object_or_404(TrainingCluster.objects.select_related("training"), pk=cluster_id)
    training = cluster.training
    if request.method == "POST":
        cluster.delete()
    return _settings_redirect(training, request)


@login_required
def update_cluster_lgas(request, cluster_id: int):
    cluster = get_object_or_404(TrainingCluster.objects.select_related("training"), pk=cluster_id)
    if request.method == "POST":
        lga_ids = request.POST.getlist("lgas")
        cluster.lgas.set(lga_ids)
    return _settings_redirect(cluster.training, request)


@login_required
def add_level(request, pk: int):
    training = get_object_or_404(Training, pk=pk)
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        days = request.POST.get("days", 1) or 1
        if name:
            TrainingLevel.objects.get_or_create(
                training=training, name=name, defaults={"days": days}
            )
    return _settings_redirect(training, request)


@login_required
def toggle_attendance(request, assignment_id: int):
    from trainings.models import TrainingAssignment
    assignment = get_object_or_404(TrainingAssignment, pk=assignment_id)
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "attended":
            assignment.attended = True
        elif action == "absent":
            assignment.attended = False
        else:
            assignment.attended = None
        assignment.attended_marked_by = request.user
        assignment.save(update_fields=["attended", "attended_marked_by"])

    # Return HTMX partial or redirect
    if request.headers.get("HX-Request"):
        from django.utils.html import format_html
        if assignment.attended is True:
            btn = f'<form method="post" action="/trainings/assignment/{assignment.pk}/attendance/" hx-post="/trainings/assignment/{assignment.pk}/attendance/" hx-target="#att-{assignment.pk}" hx-swap="innerHTML" class="inline">'
            btn += f'<input type="hidden" name="csrfmiddlewaretoken" value="{request.META.get("CSRF_COOKIE", "")}">'
            btn += '<button name="action" value="absent" class="text-[10px] px-2 py-0.5 rounded-full bg-green-50 text-green-700">Present</button></form>'
        elif assignment.attended is False:
            btn = f'<form method="post" action="/trainings/assignment/{assignment.pk}/attendance/" hx-post="/trainings/assignment/{assignment.pk}/attendance/" hx-target="#att-{assignment.pk}" hx-swap="innerHTML" class="inline">'
            btn += f'<input type="hidden" name="csrfmiddlewaretoken" value="{request.META.get("CSRF_COOKIE", "")}">'
            btn += '<button name="action" value="attended" class="text-[10px] px-2 py-0.5 rounded-full bg-red-50 text-red-600">Absent</button></form>'
        else:
            btn = f'<form method="post" action="/trainings/assignment/{assignment.pk}/attendance/" hx-post="/trainings/assignment/{assignment.pk}/attendance/" hx-target="#att-{assignment.pk}" hx-swap="innerHTML" class="inline">'
            btn += f'<input type="hidden" name="csrfmiddlewaretoken" value="{request.META.get("CSRF_COOKIE", "")}">'
            btn += '<button name="action" value="attended" class="text-[10px] px-2 py-0.5 rounded-full bg-gray-100 text-gray-500">Mark</button></form>'
        return HttpResponse(btn)

    return redirect("training_detail", pk=assignment.training_id)


@login_required
def delete_level(request, level_id: int):
    level = get_object_or_404(TrainingLevel.objects.select_related("training"), pk=level_id)
    training = level.training
    if request.method == "POST":
        level.delete()
    return _settings_redirect(training, request)


@login_required
def search_participants_for_training(request, pk: int):
    """HTMX: search participants to add to training."""
    from participants.models import Participant

    training = get_object_or_404(Training, pk=pk)
    q = request.GET.get("q", "").strip()
    results = []

    if len(q) >= 2:
        already_assigned = training.assignments.values_list("participant_id", flat=True)
        results = Participant.objects.filter(
            models.Q(first_name__icontains=q) | models.Q(last_name__icontains=q) |
            models.Q(email__icontains=q) | models.Q(phone__icontains=q)
        ).exclude(pk__in=already_assigned).select_related("channel", "state")[:10]

    return render(request, "trainings/partials/participant_search_results.html", {
        "results": results,
        "training": training,
        "q": q,
    })


@login_required
def assign_participant_to_training(request, pk: int):
    """Assign a participant to this training with role + category + cluster."""
    from participants.models import Participant
    from trainings.models import TrainingAssignment

    training = get_object_or_404(Training, pk=pk)

    if request.method == "POST":
        participant_id = request.POST.get("participant_id")
        role_id = request.POST.get("role") or None
        category_id = request.POST.get("category") or None
        cluster_id = request.POST.get("cluster") or None

        participant = get_object_or_404(Participant, pk=participant_id)

        from common.models import TrainingRole

        assignment, created = TrainingAssignment.objects.get_or_create(
            training=training,
            participant=participant,
            defaults={
                "role_id": role_id,
                "training_category_id": category_id,
                "cluster_id": cluster_id,
            },
        )
        if created:
            messages.success(request, f"{participant.full_name} added to training.")
        else:
            messages.info(request, f"{participant.full_name} is already assigned.")

    return redirect("training_detail", pk=pk)


@login_required
def remove_assignment(request, assignment_id: int):
    """Remove a participant from a training."""
    from trainings.models import TrainingAssignment

    assignment = get_object_or_404(TrainingAssignment, pk=assignment_id)
    training_pk = assignment.training_id
    name = assignment.participant.full_name

    if request.method == "POST":
        assignment.delete()
        messages.success(request, f"{name} removed from training.")

    return redirect("training_detail", pk=training_pk)


@login_required
def search_managers(request, pk: int):
    """HTMX: search users to add as training managers."""
    from accounts.models import CustomUser

    training = get_object_or_404(Training, pk=pk)
    q = request.GET.get("q", "").strip()
    results = []

    if len(q) >= 2:
        already = training.managers.values_list("pk", flat=True)
        results = CustomUser.objects.filter(
            models.Q(first_name__icontains=q) | models.Q(last_name__icontains=q) |
            models.Q(email__icontains=q) | models.Q(username__icontains=q)
        ).filter(is_active=True).exclude(pk__in=already)[:10]

    return render(request, "trainings/partials/manager_search_results.html", {
        "results": results, "training": training, "q": q,
    })


@login_required
def add_manager(request, pk: int):
    training = get_object_or_404(Training, pk=pk)
    if request.method == "POST":
        from accounts.models import CustomUser
        user_id = request.POST.get("user_id")
        user = get_object_or_404(CustomUser, pk=user_id)
        training.managers.add(user)
    return _settings_redirect(training, request)


@login_required
def remove_manager(request, pk: int, user_id: int):
    training = get_object_or_404(Training, pk=pk)
    if request.method == "POST":
        from accounts.models import CustomUser
        user = get_object_or_404(CustomUser, pk=user_id)
        training.managers.remove(user)
    return _settings_redirect(training, request)


@login_required
def search_device_managers(request, pk: int):
    from accounts.models import CustomUser
    training = get_object_or_404(Training, pk=pk)
    q = request.GET.get("q", "").strip()
    results = []
    if len(q) >= 2:
        already = training.device_managers.values_list("pk", flat=True)
        results = CustomUser.objects.filter(
            models.Q(first_name__icontains=q) | models.Q(last_name__icontains=q) |
            models.Q(email__icontains=q) | models.Q(username__icontains=q)
        ).filter(is_active=True).exclude(pk__in=already)[:10]
    return render(request, "trainings/partials/device_manager_search_results.html", {
        "results": results, "training": training, "q": q,
    })


@login_required
def add_device_manager(request, pk: int):
    training = get_object_or_404(Training, pk=pk)
    if request.method == "POST":
        from accounts.models import CustomUser
        user = get_object_or_404(CustomUser, pk=request.POST.get("user_id"))
        training.device_managers.add(user)
    return _settings_redirect(training, request)


@login_required
def remove_device_manager(request, pk: int, user_id: int):
    training = get_object_or_404(Training, pk=pk)
    if request.method == "POST":
        from accounts.models import CustomUser
        user = get_object_or_404(CustomUser, pk=user_id)
        training.device_managers.remove(user)
    return _settings_redirect(training, request)


@login_required
def toggle_devices(request, pk: int):
    training = get_object_or_404(Training, pk=pk)
    if request.method == "POST":
        training.device_management = not training.device_management
        training.save(update_fields=["device_management"])
    return redirect("training_detail", pk=pk)


@login_required
def save_training_settings(request, pk: int):
    training = get_object_or_404(Training, pk=pk)
    if request.method == "POST":
        training.device_management = "device_management" in request.POST
        training.save(update_fields=["device_management"])
    return _settings_redirect(training, request)


def _is_admin(user) -> bool:
    """Check if user is superuser or in admin group (including State Admin for exports)."""
    return user.is_superuser or user.groups.filter(
        name__in=["UNICEF Admin", "National Admin", "State Admin"]
    ).exists()


def _is_training_admin_or_manager(user) -> bool:
    """Check if user is any kind of admin or training manager."""
    if user.is_superuser:
        return True
    if user.groups.filter(name__in=["UNICEF Admin", "National Admin", "State Admin"]).exists():
        return True
    if user.managed_trainings.exists():
        return True
    return False


@login_required
def training_export_financials(request, pk: int):
    """Export training financials as CSV. Superuser/UNICEF Admin only."""
    if not _is_admin(request.user):
        messages.error(request, "You do not have permission to export financials.")
        return redirect("training_detail", pk=pk)

    training = get_object_or_404(Training, pk=pk)
    categories = training.categories.prefetch_related("levels")

    # Channel filter — state admins export only their channel
    channel_id = request.GET.get("channel")
    if not channel_id and request.user.groups.filter(name="State Admin").exists() and request.user.channel:
        channel_id = request.user.channel_id
    if channel_id:
        from django.db.models import Q
        categories = categories.filter(Q(channel_id=channel_id) | Q(channel__isnull=True, training__channel_id=channel_id))

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{training.title}_financials.csv"'

    writer = csv.writer(response)
    writer.writerow(["Category", "Channel", "DSA/Day", "Days", "Transport", "Travel Mode", "People", "Total"])

    grand_total = 0
    for cat in categories:
        cat_assignments = cat.assignments.all()
        count = cat_assignments.count()
        cat_total = sum(a.total_payment for a in cat_assignments)
        grand_total += cat_total
        writer.writerow([
            cat.name,
            str(cat.effective_channel),
            cat.dsa_rate,
            cat.total_days,
            cat.local_transport,
            cat.get_travel_mode_display(),
            count,
            cat_total,
        ])

    writer.writerow(["Grand Total", "", "", "", "", "", grand_total])
    return response


@login_required
def training_export_participants(request, pk: int):
    """Export training participants as CSV."""
    if not _is_training_admin_or_manager(request.user):
        messages.error(request, "You do not have permission to export participants.")
        return redirect("training_detail", pk=pk)

    training = get_object_or_404(Training, pk=pk)
    assignments = (
        training.assignments
        .select_related(
            "participant", "participant__channel", "participant__lga",
            "training_category", "cluster",
        )
        .order_by("training_category__name", "participant__last_name")
    )

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{training.title}_participants.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "Name", "Channel", "Category", "Cluster", "LGA",
        "Phone", "Email", "Bank", "Account Number", "Account Name", "Validated",
    ])

    for a in assignments:
        p = a.participant
        bank_name = ""
        acct_number = ""
        acct_name = ""
        validated = ""
        try:
            ba = p.bank_account
            bank_name = str(ba.bank) if ba else ""
            acct_number = ba.account_number if ba else ""
            acct_name = ba.account_name if ba else ""
            validated = "Yes" if ba and ba.is_validated else "No"
        except Exception:
            pass

        writer.writerow([
            p.full_name,
            str(p.channel) if p.channel else "",
            a.training_category.name if a.training_category else "",
            a.cluster.name if a.cluster else "",
            str(p.lga) if p.lga else "",
            p.phone,
            p.email,
            bank_name,
            acct_number,
            acct_name,
            validated,
        ])

    return response


@login_required
def training_export_provisioning(request, pk: int):
    """Export provisioning data — full participant profile for onboarding."""
    if not _is_training_admin_or_manager(request.user):
        messages.error(request, "Permission denied.")
        return redirect("training_detail", pk=pk)

    training = get_object_or_404(Training, pk=pk)
    assignments = (
        training.assignments
        .select_related(
            "participant", "participant__channel", "participant__state",
            "participant__lga", "training_category", "cluster",
        )
        .order_by("training_category__name", "participant__last_name")
    )

    # Channel filter
    channel_id = request.GET.get("channel")
    if channel_id:
        from django.db.models import Q
        assignments = assignments.filter(
            Q(training_category__channel_id=channel_id) |
            Q(training_category__channel__isnull=True, training__channel_id=channel_id)
        )

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{training.title}_provisioning.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "S/N", "Full Name", "Phone", "NIN", "Email",
        "Channel", "Organisation", "Role",
        "State", "LGA", "Category", "Operational Area",
    ])

    for i, a in enumerate(assignments, 1):
        p = a.participant
        writer.writerow([
            i,
            p.full_name,
            p.phone,
            p.nin if hasattr(p, "nin") else "",
            p.email,
            str(p.channel) if p.channel else "",
            p.health_organization,
            p.channel_role,
            str(p.state) if p.state else "",
            str(p.lga) if p.lga else "",
            a.training_category.name if a.training_category else "",
            a.cluster.name if a.cluster else "",
        ])

    return response


@login_required
def training_export_devices(request, pk: int):
    """Export training devices as CSV."""
    user = request.user
    has_access = (
        _is_training_admin_or_manager(user)
        or hasattr(user, "participant_profile")
    )
    if not has_access:
        messages.error(request, "You do not have permission to export devices.")
        return redirect("training_detail", pk=pk)

    training = get_object_or_404(Training, pk=pk)
    devices = training.devices.select_related("assigned_to", "assigned_to__lga", "state")

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{training.title}_devices.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "Serial Number", "IMEI 1", "IMEI 2", "Brand", "Model",
        "Assigned To", "Phone", "Email", "LGA",
        "Status", "Condition", "Accessories",
    ])

    for d in devices:
        p = d.assigned_to
        writer.writerow([
            d.serial_number,
            d.imei_1,
            d.imei_2,
            d.brand,
            d.model_name,
            p.full_name if p else "",
            p.phone if p else "",
            p.email if p else "",
            str(p.lga) if p and p.lga else "",
            d.get_status_display(),
            d.get_condition_display(),
            "Yes" if d.accessories_complete else "No",
        ])

    return response


@login_required
def training_exports_page(request, pk: int):
    """Export page — select what to export and filter by channel."""
    training = get_object_or_404(Training.objects.select_related("channel", "state"), pk=pk)
    from common.models import Channel
    channels = Channel.objects.filter(is_active=True)
    return render(request, "trainings/exports.html", {
        "training": training,
        "channels": channels,
    })


@login_required
def training_export_bank(request, pk: int):
    """Export bank details in UNICEF format. Leading zeros preserved. Channel-filtered."""
    training = get_object_or_404(Training, pk=pk)
    channel_id = request.GET.get("channel")
    format_type = request.GET.get("format", "bank")

    assignments = (
        training.assignments
        .select_related("participant", "participant__channel", "participant__lga", "training_category")
        .order_by("participant__last_name")
    )
    if channel_id:
        assignments = assignments.filter(participant__channel_id=channel_id)

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{training.title}_bank.csv"'
    response.write('\ufeff')

    writer = csv.writer(response)

    if format_type == "validated":
        writer.writerow(["ACCOUNT_NO", "BENEFICIARY_NAME", "NIBSS_NAME", "BANK_CODE", "BANK_NAME", "PROXY", "PROXY_NAME", "STATUS"])
        for a in assignments:
            p = a.participant
            try:
                ba = p.bank_account
                writer.writerow([
                    "'" + ba.account_number if ba else "",
                    p.full_name.upper(),
                    ba.account_name.upper() if ba and ba.account_name else "",
                    "'" + ba.bank.cbn_code if ba else "",
                    ba.bank.name.upper() if ba else "",
                    "YES" if ba and ba.is_proxy else "NO",
                    ba.proxy_name if ba and ba.is_proxy else "",
                    "VALIDATED" if ba and ba.is_validated else "NOT VALIDATED",
                ])
            except Exception:
                writer.writerow(["", p.full_name.upper(), "", "", "", "NO", "", "NO ACCOUNT"])
    else:
        writer.writerow(["S/N", "ACCOUNT_NO", "BENEFICIARY_NAME", "AMOUNT", "SENDER", "BANK_NAME", "CBN_CODE", "UNICEF_BANK_ID", "CHANNEL", "LGA", "PROXY", "PROXY_NAME", "DESCRIPTION"])
        serial = 0
        for a in assignments:
            p = a.participant
            try:
                ba = p.bank_account
                if not ba:
                    continue
                serial += 1
                writer.writerow([
                    serial,
                    "'" + ba.account_number,
                    p.full_name.upper(),
                    a.total_payment,
                    "UNICEF",
                    ba.bank.name.upper(),
                    "'" + ba.bank.cbn_code,
                    "'" + ba.bank.unicef_code,
                    str(p.channel) if p.channel else "",
                    str(p.lga) if p.lga else "",
                    "YES" if ba.is_proxy else "NO",
                    ba.proxy_name if ba.is_proxy else "",
                    training.title,
                ])
            except Exception:
                continue

    return response


@login_required
def training_export_bank_excel(request, pk: int):
    """Export bank details as Excel in UNICEF payment format."""
    import openpyxl
    from io import BytesIO

    training = get_object_or_404(Training, pk=pk)
    channel_id = request.GET.get("channel")

    assignments = (
        training.assignments
        .select_related("participant", "participant__channel", "participant__lga", "training_category")
        .order_by("participant__last_name")
    )
    if channel_id:
        assignments = assignments.filter(participant__channel_id=channel_id)

    wb = openpyxl.Workbook()

    # Sheet 1: Payment template
    ws1 = wb.active
    ws1.title = "Payment Template"
    ws1.append(["S/N", "ACCOUNT_NO", "BENEFICIARY_NAME", "AMOUNT", "SENDER", "BANK_NAME", "CBN_CODE", "UNICEF_BANK_ID", "CHANNEL", "LGA", "PROXY", "PROXY_NAME", "DESCRIPTION"])
    serial = 0
    for a in assignments:
        p = a.participant
        try:
            ba = p.bank_account
            if not ba:
                continue
            serial += 1
            ws1.append([
                serial,
                ba.account_number,
                p.full_name.upper(),
                float(a.total_payment),
                "UNICEF",
                ba.bank.name.upper(),
                ba.bank.cbn_code,
                ba.bank.unicef_code,
                str(p.channel) if p.channel else "",
                str(p.lga) if p.lga else "",
                "YES" if ba.is_proxy else "NO",
                ba.proxy_name if ba.is_proxy else "",
                training.title,
            ])
        except Exception:
            continue

    # Sheet 2: Validation report
    ws2 = wb.create_sheet("Validation Report")
    ws2.append(["ACCOUNT_NO", "BENEFICIARY_NAME", "NIBSS_NAME", "BANK_CODE", "BANK_NAME", "PROXY", "PROXY_NAME", "STATUS"])
    for a in assignments:
        p = a.participant
        try:
            ba = p.bank_account
            ws2.append([
                ba.account_number if ba else "",
                p.full_name.upper(),
                ba.account_name.upper() if ba and ba.account_name else "",
                ba.bank.cbn_code if ba else "",
                ba.bank.name.upper() if ba else "",
                "YES" if ba and ba.is_proxy else "NO",
                ba.proxy_name if ba and ba.is_proxy else "",
                "VALIDATED" if ba and ba.is_validated else "NOT VALIDATED",
            ])
        except Exception:
            ws2.append(["", p.full_name.upper(), "", "", "", "NO", "", "NO ACCOUNT"])

    # Sheet 3: Payment Schedule (financials breakdown)
    ws3 = wb.create_sheet("Payment Schedule")
    ws3.append([training.title.upper()])
    ws3.append([f"State: {training.state}"])
    ws3.append([f"Implementing Partner: {training.implementing_partner}"])
    ws3.append([f"Responsible Officer: {training.responsible_officer}"])
    ws3.append([])
    ws3.append([
        "S/N", "BENEFICIARY NAME", "CATEGORY", "CHANNEL", "LGA",
        "DSA RATE", "TRAINING DAYS", "DSA TOTAL",
        "LOCAL TRANSPORT", "TRANSPORT DAYS", "TRANSPORT TOTAL",
        "TERMINALS", "ROAD MILEAGE", "AIR CLAIM",
        "GRAND TOTAL", "PROXY", "BANK", "ACCOUNT NO", "CBN CODE", "UNICEF BANK ID"
    ])
    serial = 0
    grand = 0
    for a in assignments:
        p = a.participant
        serial += 1
        total = float(a.total_payment)
        grand += total
        bank_name = ""
        acct_no = ""
        cbn = ""
        uid = ""
        is_proxy = "NO"
        try:
            ba = p.bank_account
            if ba:
                bank_name = ba.bank.name.upper()
                acct_no = ba.account_number
                cbn = ba.bank.cbn_code
                uid = ba.bank.unicef_code
                if ba.is_proxy:
                    is_proxy = "YES"
        except Exception:
            pass

        cat = a.training_category
        days = cat.total_days if cat else 0
        dsa_rate = float(cat.dsa_rate) if cat else 0
        transport = float(cat.local_transport) if cat else 0

        ws3.append([
            serial,
            p.full_name.upper(),
            cat.name if cat else "",
            str(p.channel) if p.channel else "",
            str(p.lga) if p.lga else "",
            dsa_rate,
            days,
            float(a.dsa_amount),
            transport,
            days,
            float(a.transport_amount),
            float(a.terminals_amount),
            float(a.outbound_travel_cost + a.return_travel_cost) if hasattr(a, 'outbound_travel_cost') else 0,
            0,
            total,
            is_proxy,
            bank_name,
            acct_no,
            cbn,
            uid,
        ])
    ws3.append([])
    ws3.append(["", "GRAND TOTAL", "", "", "", "", "", "", "", "", "", "", "", "", grand])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{training.title}_bank.xlsx"'
    return response


@login_required
def training_export_financials_excel(request, pk: int):
    """Export financials as Excel."""
    if not _is_admin(request.user):
        return redirect("training_detail", pk=pk)

    import openpyxl
    from io import BytesIO

    training = get_object_or_404(Training.objects.select_related("state"), pk=pk)
    categories = training.categories.prefetch_related("levels")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment Schedule"

    # Header info
    ws.append([training.title.upper()])
    ws.append([f"State: {training.state}"])
    ws.append([f"Implementing Partner: {training.implementing_partner}"])
    ws.append([f"Responsible Officer: {training.responsible_officer}"])
    ws.append([])

    # Table header
    ws.append(["S/N", "CATEGORY", "DSA/DAY", "DAYS", "DSA TOTAL", "TRANSPORT", "TRAVEL", "TERMINALS", "PEOPLE", "TOTAL"])

    grand_total = 0
    serial = 0
    for cat in categories:
        cat_assignments = cat.assignments.all()
        count = cat_assignments.count()
        cat_total = sum(a.total_payment for a in cat_assignments)
        grand_total += cat_total
        serial += 1
        ws.append([
            serial,
            cat.name,
            float(cat.dsa_rate),
            cat.total_days,
            float(cat.dsa_total),
            float(cat.local_transport),
            cat.get_travel_mode_display(),
            "",
            count,
            float(cat_total),
        ])

    ws.append([])
    ws.append(["", "GRAND TOTAL", "", "", "", "", "", "", "", float(grand_total)])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{training.title}_financials.xlsx"'
    return response


@login_required
def send_edit_links(request, pk: int):
    """Send profile edit links via email to participants without LGA in this training."""
    from django.conf import settings as django_settings
    from django.core.mail import send_mail

    training = get_object_or_404(Training, pk=pk)
    if request.method != "POST":
        return redirect("training_detail", pk=pk)

    base_url = getattr(django_settings, "BASE_URL", "https://tma.worksiapps.com")
    assignments = training.assignments.select_related("participant").filter(
        participant__lga__isnull=True,
    )

    sent = 0
    no_email = 0
    for a in assignments:
        p = a.participant
        if not p.email:
            no_email += 1
            continue
        edit_link = f"{base_url}{p.edit_url}"
        try:
            send_mail(
                "UNICEF TMA - Please Update Your LGA",
                f"Dear {p.full_name},\n\n"
                f"Your profile for the training \"{training.title}\" is missing your LGA (Local Government Area).\n\n"
                f"Please click the link below to update your profile:\n"
                f"{edit_link}\n\n"
                f"This is important for your cluster assignment and payment processing.\n\n"
                f"---\nUNICEF Training Management Application",
                None,
                [p.email],
                fail_silently=True,
            )
            sent += 1
        except Exception:
            pass

    msg = f"Sent {sent} email(s) with profile edit links."
    if no_email:
        msg += f" {no_email} participant(s) have no email address."
    messages.success(request, msg)
    return redirect("training_detail", pk=pk)


@login_required
def dm_send_login(request, pk: int, participant_id: int):
    """Reset a device manager's password and email them login + guide."""
    import secrets
    from django.conf import settings as s
    from django.core.mail import send_mail
    from participants.models import Participant

    training = get_object_or_404(Training, pk=pk)
    p = get_object_or_404(Participant, pk=participant_id)

    if request.method != "POST" or not p.user:
        return redirect("training_detail", pk=pk)

    new_pw = secrets.token_urlsafe(8)
    p.user.set_password(new_pw)
    p.user.save()

    if p.email:
        base = getattr(s, "BASE_URL", "https://tma.worksiapps.com")
        try:
            send_mail(
                "UNICEF TMA - Your Login Details",
                f"Dear {p.full_name},\n\n"
                f"Your login for the UNICEF Training Management Application:\n\n"
                f"Website: {base}\n"
                f"Username: {p.user.username}\n"
                f"Password: {new_pw}\n\n"
                f"Training: {training.title}\n"
                f"LGA: {p.lga or 'Not set'}\n\n"
                f"Please change your password after logging in.\n\n"
                f"---\nUNICEF Training Management Application",
                None, [p.email], fail_silently=True,
            )
            messages.success(request, f"Login sent to {p.full_name} ({p.email}). Password: {new_pw}")
        except Exception:
            messages.warning(request, f"Password reset to {new_pw} but email failed.")
    else:
        messages.success(request, f"Password reset for {p.full_name}. New password: {new_pw} (no email on file)")

    return redirect("training_detail", pk=pk)


@login_required
def dm_send_all(request, pk: int):
    """Send login details to ALL device managers in this training."""
    import secrets
    from django.conf import settings as s
    from django.core.mail import send_mail

    training = get_object_or_404(Training, pk=pk)
    if request.method != "POST":
        return redirect("training_detail", pk=pk)

    dm_assignments = training.assignments.filter(
        training_category__is_device_manager=True,
        participant__user__isnull=False,
    ).select_related("participant", "participant__user", "participant__lga")

    base = getattr(s, "BASE_URL", "https://tma.worksiapps.com")
    sent = 0
    for a in dm_assignments:
        p = a.participant
        if not p.email:
            continue
        new_pw = secrets.token_urlsafe(8)
        p.user.set_password(new_pw)
        p.user.save()
        try:
            send_mail(
                "UNICEF TMA - Your Login Details",
                f"Dear {p.full_name},\n\n"
                f"Your login for the UNICEF Training Management Application:\n\n"
                f"Website: {base}\n"
                f"Username: {p.user.username}\n"
                f"Password: {new_pw}\n\n"
                f"Training: {training.title}\n"
                f"LGA: {p.lga or 'Not set'}\n\n"
                f"Please change your password after logging in.\n\n"
                f"---\nUNICEF Training Management Application",
                None, [p.email], fail_silently=True,
            )
            sent += 1
        except Exception:
            pass

    messages.success(request, f"Sent login details to {sent} device manager(s).")
    return redirect("training_detail", pk=pk)


@login_required
def dm_toggle(request, pk: int, user_id: int):
    """Enable/disable a device manager's login."""
    from accounts.models import CustomUser
    training = get_object_or_404(Training, pk=pk)
    user = get_object_or_404(CustomUser, pk=user_id)
    if request.method == "POST":
        user.is_active = not user.is_active
        user.save(update_fields=["is_active"])
        status = "enabled" if user.is_active else "disabled"
        messages.success(request, f"{user.get_full_name() or user.username} {status}.")
    return redirect("training_detail", pk=pk)


@login_required
def dm_create_login(request, pk: int, participant_id: int):
    """Create a login account for a device manager who doesn't have one."""
    import secrets
    from accounts.models import CustomUser
    from django.conf import settings as s
    from django.core.mail import send_mail
    from participants.models import Participant

    training = get_object_or_404(Training, pk=pk)
    p = get_object_or_404(Participant, pk=participant_id)
    if request.method != "POST" or p.user:
        return redirect("training_detail", pk=pk)

    username = p.phone or p.email or f"dm_{p.pk}"
    if CustomUser.objects.filter(username=username).exists():
        username = f"dm_{p.pk}"

    new_pw = secrets.token_urlsafe(8)
    user = CustomUser.objects.create_user(
        username=username, password=new_pw,
        first_name=p.first_name, last_name=p.last_name,
        email=p.email, state=p.state, lga=p.lga,
    )
    p.user = user
    p.save(update_fields=["user"])

    if p.email:
        base = getattr(s, "BASE_URL", "https://tma.worksiapps.com")
        try:
            send_mail(
                "UNICEF TMA - Your Login Details",
                f"Dear {p.full_name},\n\n"
                f"A login account has been created for you:\n\n"
                f"Website: {base}\n"
                f"Username: {username}\n"
                f"Password: {new_pw}\n\n"
                f"Training: {training.title}\n\n"
                f"---\nUNICEF Training Management Application",
                None, [p.email], fail_silently=True,
            )
        except Exception:
            pass

    messages.success(request, f"Login created for {p.full_name}. Username: {username}, Password: {new_pw}")
    return redirect("training_detail", pk=pk)
